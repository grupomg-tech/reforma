from decimal import Decimal
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt

from apps.empresa.models import Empresa
from apps.dashboards.models import ProdutoSaidaAPI, AjusteManualICMS


@login_required
def index(request):
    return render(request, 'dashboards/index.html')


@login_required
def relatorio_fiscal(request):
    from django.db.models import Sum
    from decimal import Decimal
    from apps.sped.models import Registro0000, Registro0200, RegistroC100, RegistroC170
    
    empresas = Empresa.objects.filter(ativo=True)
    
    # Parâmetros do filtro
    empresa_id = request.GET.get('empresa')
    periodo_inicial = request.GET.get('periodo_inicial')
    periodo_final = request.GET.get('periodo_final')
    aliquota_ibs = Decimal(request.GET.get('aliquota_ibs', '18.5') or '18.5')
    aliquota_cbs = Decimal(request.GET.get('aliquota_cbs', '8.5') or '8.5')
    aliquota_is = Decimal(request.GET.get('aliquota_is', '0') or '0')
    agrupar_filiais = request.GET.get('filiais') == '1'
    
    # Contexto base
    context = {
        'empresas': empresas,
        'periodos': [],
    }
    
    # Se empresa selecionada, buscar períodos disponíveis
    if empresa_id:
        periodos = Registro0000.objects.filter(
            empresa_id=empresa_id,
            processado=True
        ).values_list('periodo', flat=True).distinct().order_by('periodo')
        context['periodos'] = [p.strftime('%Y-%m') for p in periodos]
    
    # Se filtro completo, buscar dados
    if empresa_id and periodo_inicial and periodo_final:
        # Buscar registros 0000 do período
        registros_0000 = Registro0000.objects.filter(
            empresa_id=empresa_id,
            processado=True,
            periodo__gte=periodo_inicial + '-01',
            periodo__lte=periodo_final + '-28'
        )
        
        if agrupar_filiais:
            empresa = Empresa.objects.get(id=empresa_id)
            cnpj_base = empresa.cnpj_cpf[:8] if empresa.cnpj_cpf else ''
            empresas_grupo = Empresa.objects.filter(cnpj_cpf__startswith=cnpj_base)
            registros_0000 = Registro0000.objects.filter(
                empresa__in=empresas_grupo,
                processado=True,
                periodo__gte=periodo_inicial + '-01',
                periodo__lte=periodo_final + '-28'
            )
        
        # Buscar documentos C100 (0=Entrada, 1=Saída)
        documentos_entrada = RegistroC100.objects.filter(registro_0000__in=registros_0000, ind_oper='0')
        documentos_saida = RegistroC100.objects.filter(registro_0000__in=registros_0000, ind_oper='1')
        
        # Buscar itens C170
        itens_entrada = RegistroC170.objects.filter(registro_c100__in=documentos_entrada).select_related('registro_c100')
        itens_saida = RegistroC170.objects.filter(registro_c100__in=documentos_saida).select_related('registro_c100')
        
        # Buscar dados dos itens (0200) para NCM e descrição
        itens_0200 = {}
        for reg in registros_0000:
            for item in Registro0200.objects.filter(registro_0000=reg):
                itens_0200[item.cod_item] = {
                    'descricao': item.descr_item,
                    'ncm': item.cod_ncm,
                }
        
        produtos_entradas_dict = {}
        for item in itens_entrada:
            cod = item.cod_item
            if cod not in produtos_entradas_dict:
                info_item = itens_0200.get(cod, {})
                produtos_entradas_dict[cod] = {
                    'codigo': cod,
                    'descricao': info_item.get('descricao', item.descr_compl or cod),
                    'ncm': item.cod_ncm or info_item.get('ncm', ''),
                    'cfop': item.cfop,
                    'quantidade': Decimal('0'),
                    'valor_total': Decimal('0'),
                    'icms': Decimal('0'),
                    'icms_st': Decimal('0'),
                    'ipi': Decimal('0'),
                    'pis': Decimal('0'),
                    'cofins': Decimal('0'),
                    'vl_bc_icms': Decimal('0'),
                    'vl_bc_pis': Decimal('0'),
                    'vl_bc_cofins': Decimal('0'),
                }
            produtos_entradas_dict[cod]['quantidade'] += item.qtd or Decimal('0')
            produtos_entradas_dict[cod]['valor_total'] += item.vl_item or Decimal('0')
            produtos_entradas_dict[cod]['icms'] += item.vl_icms or Decimal('0')
            produtos_entradas_dict[cod]['icms_st'] += getattr(item, 'vl_icms_st', Decimal('0')) or Decimal('0')
            produtos_entradas_dict[cod]['ipi'] += getattr(item, 'vl_ipi', Decimal('0')) or Decimal('0')
            produtos_entradas_dict[cod]['pis'] += item.vl_pis or Decimal('0')
            produtos_entradas_dict[cod]['cofins'] += item.vl_cofins or Decimal('0')
            produtos_entradas_dict[cod]['vl_bc_icms'] += item.vl_bc_icms or Decimal('0')
            produtos_entradas_dict[cod]['vl_bc_pis'] += item.vl_bc_pis or Decimal('0')
            produtos_entradas_dict[cod]['vl_bc_cofins'] += item.vl_bc_cofins or Decimal('0')
        
# ========================================
        # NCMs com redução de 60% (LC 214/2025 - Anexo XVII)
        # Construção civil e materiais correlatos
        # ========================================
        NCM_REDUCAO_60 = {
            '25222000',  # Gesso
            '38244000',  # Impermeabilizantes / produtos químicos p/ construção
        }
        
        from decimal import ROUND_HALF_UP
        
# Calcular IBS/CBS e formatar campos para entradas
        for cod, prod in produtos_entradas_dict.items():
            # Valores calculados
            total_tributos = prod['icms'] + prod['icms_st'] + prod['ipi'] + prod['pis'] + prod['cofins']
            valor_liquido = prod['valor_total'] - total_tributos
            
            # Verificar se o NCM tem redução
            ncm_prod = str(prod['ncm']).strip()
            if ncm_prod in NCM_REDUCAO_60:
                perc_reducao = Decimal('60')
            else:
                perc_reducao = Decimal('0')
            
            aliq_total = aliquota_ibs + aliquota_cbs + aliquota_is
            aliq_efetiva = aliq_total * (Decimal('1') - perc_reducao / Decimal('100'))
            ibs_cbs = (valor_liquido * aliq_efetiva / Decimal('100')).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            total_reforma = valor_liquido + ibs_cbs
            dif_total = ibs_cbs - total_tributos
            
            # Alíquotas calculadas a partir das bases agregadas
            aliq_icms_calc = (prod['icms'] / prod['vl_bc_icms'] * 100).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP) if prod['vl_bc_icms'] > 0 else Decimal('0')
            aliq_pis_calc = (prod['pis'] / prod['vl_bc_pis'] * 100).quantize(Decimal('0.0001'), rounding=ROUND_HALF_UP) if prod['vl_bc_pis'] > 0 else Decimal('0')
            aliq_cofins_calc = (prod['cofins'] / prod['vl_bc_cofins'] * 100).quantize(Decimal('0.0001'), rounding=ROUND_HALF_UP) if prod['vl_bc_cofins'] > 0 else Decimal('0')
            
            # Valores unitários
            qtd = prod['quantidade'] if prod['quantidade'] > 0 else Decimal('1')
            valor_bruto_unit = (prod['valor_total'] / qtd).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            valor_liq_unit = (valor_liquido / qtd).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            ibs_cbs_unit = (ibs_cbs / qtd).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            total_reforma_unit = (total_reforma / qtd).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            dif_unit = (dif_total / qtd).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            
            # Atualiza produto com valores calculados
            prod['ibs_cbs'] = ibs_cbs
            prod['valor_liquido'] = valor_liquido
            prod['total_tributos'] = total_tributos
            prod['total_reforma'] = total_reforma
            prod['dif_total'] = dif_total
            prod['valor_bruto_unit'] = valor_bruto_unit
            prod['valor_liq_unit'] = valor_liq_unit
            prod['ibs_cbs_unit'] = ibs_cbs_unit
            prod['total_reforma_unit'] = total_reforma_unit
            prod['dif_unit'] = dif_unit
            prod['aliq_ibs_cbs'] = aliq_efetiva
            prod['perc_reducao'] = perc_reducao
            prod['aliq_icms'] = aliq_icms_calc
            prod['aliq_pis'] = aliq_pis_calc
            prod['aliq_cofins'] = aliq_cofins_calc
            prod['aliq_icms'] = aliq_icms_calc
            prod['aliq_pis'] = aliq_pis_calc
            prod['aliq_cofins'] = aliq_cofins_calc
            
# Formatar para exibição (padrão brasileiro)
            prod['quantidade_fmt'] = f"{prod['quantidade']:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
            prod['perc_reducao_fmt'] = f"{perc_reducao:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
            prod['aliq_ibs_cbs_fmt'] = f"{aliq_efetiva:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
            prod['aliq_icms_fmt'] = f"{aliq_icms_calc:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
            prod['aliq_pis_fmt'] = f"{aliq_pis_calc:,.4f}".replace(',', 'X').replace('.', ',').replace('X', '.')
            prod['aliq_cofins_fmt'] = f"{aliq_cofins_calc:,.4f}".replace(',', 'X').replace('.', ',').replace('X', '.')
            prod['valor_bruto_unit_fmt'] = f"{valor_bruto_unit:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
            prod['valor_bruto_fmt'] = f"{prod['valor_total']:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
            prod['icms_fmt'] = f"{prod['icms']:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
            prod['icms_st_fmt'] = f"{prod['icms_st']:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
            prod['ipi_fmt'] = f"{prod['ipi']:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
            prod['pis_fmt'] = f"{prod['pis']:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
            prod['cofins_fmt'] = f"{prod['cofins']:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
            prod['total_tributos_fmt'] = f"{total_tributos:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
            prod['valor_liq_unit_fmt'] = f"{valor_liq_unit:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
            prod['valor_liq_fmt'] = f"{valor_liquido:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
            prod['ibs_cbs_unit_fmt'] = f"{ibs_cbs_unit:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
            prod['ibs_cbs_fmt'] = f"{ibs_cbs:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
            prod['total_reforma_unit_fmt'] = f"{total_reforma_unit:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
            prod['total_reforma_fmt'] = f"{total_reforma:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
            prod['dif_unit_fmt'] = f"{dif_unit:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
            prod['dif_total_fmt'] = f"{dif_total:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
        
        produtos_entradas = sorted(produtos_entradas_dict.values(), key=lambda x: (-x['valor_total'], x['codigo']))
        
        # ========================================
        # AGREGAR FORNECEDORES DE ENTRADA
        # ========================================
        from apps.sped.models import Registro0150
        
        # Buscar participantes (fornecedores) dos registros
        participantes_dict = {}
        for reg in registros_0000:
            for part in Registro0150.objects.filter(registro_0000=reg):
                if part.cod_part not in participantes_dict:
                    # Determinar regime tributário
                    if part.optante_mei:
                        regime = 'MEI'
                    elif part.optante_simples:
                        regime = 'SIMPLES NACIONAL'
                    elif part.cnpj:
                        if not part.consultado:
                            regime = 'NÃO CONSULTADO'
                        else:
                            regime = 'LUCRO REAL/PRESUMIDO'
                    else:
                        regime = 'PESSOA FÍSICA' if part.cpf else '-'
                    
                    participantes_dict[part.cod_part] = {
                        'codigo': part.cod_part,
                        'cnpj_cpf': part.cnpj or part.cpf or '-',
                        'nome': part.nome,
                        'regime': regime,
                        'uf': part.uf,
                    }
        
        # Agregar valores por fornecedor (documentos de entrada)
        fornecedores_dict = {}
        for doc in documentos_entrada:
            cod_part = doc.cod_part
            if not cod_part:
                continue
            
            if cod_part not in fornecedores_dict:
                info_part = participantes_dict.get(cod_part, {})
                fornecedores_dict[cod_part] = {
                    'codigo': cod_part,
                    'cnpj_cpf': info_part.get('cnpj_cpf', '-'),
                    'nome': info_part.get('nome', 'Não identificado'),
                    'regime': info_part.get('regime', '-'),
                    'uf': info_part.get('uf', '-'),
                    'valor_bruto': Decimal('0'),
                    'icms': Decimal('0'),
                    'icms_st': Decimal('0'),
                    'ipi': Decimal('0'),
                    'iss': Decimal('0'),
                    'pis': Decimal('0'),
                    'cofins': Decimal('0'),
                }
            
            fornecedores_dict[cod_part]['valor_bruto'] += doc.vl_doc or Decimal('0')
            fornecedores_dict[cod_part]['icms'] += doc.vl_icms or Decimal('0')
            fornecedores_dict[cod_part]['pis'] += doc.vl_pis or Decimal('0')
            fornecedores_dict[cod_part]['cofins'] += doc.vl_cofins or Decimal('0')
        
        # Calcular tributos totais e valores reforma para fornecedores
        for cod, forn in fornecedores_dict.items():
            tributos = forn['icms'] + forn['icms_st'] + forn['ipi'] + forn['iss'] + forn['pis'] + forn['cofins']
            liquido = forn['valor_bruto'] - tributos
            aliq_total = aliquota_ibs + aliquota_cbs + aliquota_is
            ibs_cbs_efetivo = (liquido * aliq_total / Decimal('100')).quantize(Decimal('0.01'))
            total_reforma = liquido + ibs_cbs_efetivo
            
            forn['tributos'] = tributos
            forn['liquido'] = liquido
            forn['ibs_cbs_efetivo'] = ibs_cbs_efetivo
            forn['total_reforma'] = total_reforma
        
        fornecedores_entradas = sorted(fornecedores_dict.values(), key=lambda x: x['valor_bruto'], reverse=True)
        
        # Totais fornecedores
        total_fornecedores_valor_bruto = sum(f['valor_bruto'] for f in fornecedores_entradas)
        total_fornecedores_icms = sum(f['icms'] for f in fornecedores_entradas)
        total_fornecedores_icms_st = sum(f['icms_st'] for f in fornecedores_entradas)
        total_fornecedores_ipi = sum(f['ipi'] for f in fornecedores_entradas)
        total_fornecedores_iss = sum(f['iss'] for f in fornecedores_entradas)
        total_fornecedores_pis = sum(f['pis'] for f in fornecedores_entradas)
        total_fornecedores_cofins = sum(f['cofins'] for f in fornecedores_entradas)
        total_fornecedores_tributos = sum(f['tributos'] for f in fornecedores_entradas)
        total_fornecedores_liquido = sum(f['liquido'] for f in fornecedores_entradas)
        total_fornecedores_ibs_cbs = sum(f['ibs_cbs_efetivo'] for f in fornecedores_entradas)
        total_fornecedores_reforma = sum(f['total_reforma'] for f in fornecedores_entradas)
        
        # ========================================
        # AGREGAR CFOPs DE ENTRADA
        # ========================================
        cfops_entrada_dict = {}
        for item in itens_entrada:
            cfop_code = item.cfop or ''
            if not cfop_code:
                continue
            
            if cfop_code not in cfops_entrada_dict:
                cfops_entrada_dict[cfop_code] = {
                    'cfop': cfop_code,
                    'valor_bruto': Decimal('0'),
                    'icms': Decimal('0'),
                    'icms_st': Decimal('0'),
                    'ipi': Decimal('0'),
                    'pis': Decimal('0'),
                    'cofins': Decimal('0'),
                }
            
            cfops_entrada_dict[cfop_code]['valor_bruto'] += item.vl_item or Decimal('0')
            cfops_entrada_dict[cfop_code]['icms'] += item.vl_icms or Decimal('0')
            cfops_entrada_dict[cfop_code]['icms_st'] += getattr(item, 'vl_icms_st', Decimal('0')) or Decimal('0')
            cfops_entrada_dict[cfop_code]['ipi'] += getattr(item, 'vl_ipi', Decimal('0')) or Decimal('0')
            cfops_entrada_dict[cfop_code]['pis'] += item.vl_pis or Decimal('0')
            cfops_entrada_dict[cfop_code]['cofins'] += item.vl_cofins or Decimal('0')
        
        # Calcular tributos e líquido para cada CFOP
        for cfop_code, cfop_data in cfops_entrada_dict.items():
            tributos = (cfop_data['icms'] + cfop_data['icms_st'] + cfop_data['ipi']
                       + cfop_data['pis'] + cfop_data['cofins'])
            cfop_data['tributos'] = tributos
            cfop_data['liquido'] = cfop_data['valor_bruto'] - tributos
        
        cfops_entradas = sorted(cfops_entrada_dict.values(), key=lambda x: x['cfop'])
        
        # Totais CFOPs
        total_cfops_valor_bruto = sum(c['valor_bruto'] for c in cfops_entradas)
        total_cfops_icms = sum(c['icms'] for c in cfops_entradas)
        total_cfops_icms_st = sum(c['icms_st'] for c in cfops_entradas)
        total_cfops_ipi = sum(c['ipi'] for c in cfops_entradas)
        total_cfops_pis = sum(c['pis'] for c in cfops_entradas)
        total_cfops_cofins = sum(c['cofins'] for c in cfops_entradas)
        total_cfops_tributos = sum(c['tributos'] for c in cfops_entradas)
        total_cfops_liquido = sum(c['liquido'] for c in cfops_entradas)
        
        # ========================================
        # AGREGAR UFs DE ENTRADA
        # ========================================
        ufs_entrada_dict = {}
        for doc in documentos_entrada:
            cod_part = doc.cod_part
            info_part = participantes_dict.get(cod_part, {})
            uf_code = info_part.get('uf', '') or ''
            if not uf_code:
                uf_code = 'N/I'
            
            if uf_code not in ufs_entrada_dict:
                ufs_entrada_dict[uf_code] = {
                    'uf': uf_code,
                    'valor_bruto': Decimal('0'),
                    'icms': Decimal('0'),
                    'icms_st': Decimal('0'),
                    'ipi': Decimal('0'),
                    'iss': Decimal('0'),
                    'pis': Decimal('0'),
                    'cofins': Decimal('0'),
                }
            
            ufs_entrada_dict[uf_code]['valor_bruto'] += doc.vl_doc or Decimal('0')
            ufs_entrada_dict[uf_code]['icms'] += doc.vl_icms or Decimal('0')
            ufs_entrada_dict[uf_code]['pis'] += doc.vl_pis or Decimal('0')
            ufs_entrada_dict[uf_code]['cofins'] += doc.vl_cofins or Decimal('0')
        
        # Calcular tributos e líquido para cada UF
        for uf_code, uf_data in ufs_entrada_dict.items():
            tributos = (uf_data['icms'] + uf_data['icms_st'] + uf_data['ipi']
                       + uf_data['iss'] + uf_data['pis'] + uf_data['cofins'])
            uf_data['tributos'] = tributos
            uf_data['liquido'] = uf_data['valor_bruto'] - tributos
        
        ufs_entradas = sorted(ufs_entrada_dict.values(), key=lambda x: x['uf'])
        
        # Totais UFs
        total_ufs_valor_bruto = sum(u['valor_bruto'] for u in ufs_entradas)
        total_ufs_icms = sum(u['icms'] for u in ufs_entradas)
        total_ufs_icms_st = sum(u['icms_st'] for u in ufs_entradas)
        total_ufs_ipi = sum(u['ipi'] for u in ufs_entradas)
        total_ufs_iss = sum(u['iss'] for u in ufs_entradas)
        total_ufs_pis = sum(u['pis'] for u in ufs_entradas)
        total_ufs_cofins = sum(u['cofins'] for u in ufs_entradas)
        total_ufs_tributos = sum(u['tributos'] for u in ufs_entradas)
        total_ufs_liquido = sum(u['liquido'] for u in ufs_entradas)
        
        # ========================================
        # AGREGAR AJUSTES APURAÇÃO ICMS (E111)
        # ========================================
        from apps.sped.parser import parse_sped_file as parse_sped
        
        TIPOS_AJUSTE = {
            '0': '0 – Outros débitos',
            '1': '1 – Estorno de créditos',
            '2': '2 – Outros créditos',
            '3': '3 – Estorno de débitos',
            '4': '4 – Deduções do imposto apurado',
            '5': '5 – Débito especial',
        }
        
        ajustes_icms_list = []
        for reg in registros_0000:
            try:
                arquivo = reg.arquivo_original
                arquivo.open('rb')
                conteudo_sped = arquivo.read()
                arquivo.close()
                dados_sped = parse_sped(conteudo_sped)
                for aj in dados_sped.get('E111', []):
                    cod = aj.get('cod_aj_apur', '')
                    tipo_char = cod[2] if len(cod) > 2 else ''
                    descricao = aj.get('descr_compl_aj', '') or TIPOS_AJUSTE.get(tipo_char, tipo_char)
                    if not descricao.strip():
                        descricao = TIPOS_AJUSTE.get(tipo_char, f'Tipo {tipo_char}')
                    ajustes_icms_list.append({
                        'codigo': cod,
                        'descricao': descricao,
                        'valor': aj.get('vl_aj_apur', Decimal('0')),
                    })
            except Exception as e:
                logger.error(f"Erro ao ler ajustes E111 do registro {reg.id}: {e}") if 'logger' in dir() else None
        
        # Agrupar por código de ajuste
        ajustes_icms_dict = {}
        for aj in ajustes_icms_list:
            cod = aj['codigo']
            if cod not in ajustes_icms_dict:
                ajustes_icms_dict[cod] = {
                    'codigo': cod,
                    'descricao': aj['descricao'],
                    'valor': Decimal('0'),
                }
            ajustes_icms_dict[cod]['valor'] += aj['valor']
        
        ajustes_icms = sorted(ajustes_icms_dict.values(), key=lambda x: x['codigo'])
        total_ajustes_icms = sum(a['valor'] for a in ajustes_icms)
        
        # ========================================
        # AJUSTES MANUAIS ICMS
        # ========================================
        ajustes_manuais_icms = AjusteManualICMS.objects.filter(
            empresa_id=empresa_id,
            periodo_inicial=periodo_inicial,
            periodo_final=periodo_final
        ).order_by('codigo')
        total_ajustes_manuais_icms = sum(a.valor for a in ajustes_manuais_icms)
        
# Agregar produtos de SAÍDA
        produtos_saidas_dict = {}
        for item in itens_saida:
            cod = item.cod_item
            if cod not in produtos_saidas_dict:
                info_item = itens_0200.get(cod, {})
                produtos_saidas_dict[cod] = {
                    'codigo': cod,
                    'descricao': info_item.get('descricao', item.descr_compl or cod),
                    'ncm': item.cod_ncm or info_item.get('ncm', ''),
                    'cfop': item.cfop,
                    'quantidade': Decimal('0'),
                    'valor_total': Decimal('0'),
                    'icms': Decimal('0'),
                    'icms_st': Decimal('0'),
                    'ipi': Decimal('0'),
                    'pis': Decimal('0'),
                    'cofins': Decimal('0'),
                }
            produtos_saidas_dict[cod]['quantidade'] += item.qtd or Decimal('0')
            produtos_saidas_dict[cod]['valor_total'] += item.vl_item or Decimal('0')
            produtos_saidas_dict[cod]['icms'] += item.vl_icms or Decimal('0')
            produtos_saidas_dict[cod]['icms_st'] += getattr(item, 'vl_icms_st', Decimal('0')) or Decimal('0')
            produtos_saidas_dict[cod]['ipi'] += getattr(item, 'vl_ipi', Decimal('0')) or Decimal('0')
            produtos_saidas_dict[cod]['pis'] += item.vl_pis or Decimal('0')
            produtos_saidas_dict[cod]['cofins'] += item.vl_cofins or Decimal('0')
        
# Calcular IBS/CBS e formatar campos para saídas
        for cod, prod in produtos_saidas_dict.items():
            total_tributos = prod['icms'] + prod['icms_st'] + prod['ipi'] + prod['pis'] + prod['cofins']
            valor_liquido = prod['valor_total'] - total_tributos
            aliq_total = aliquota_ibs + aliquota_cbs + aliquota_is
            ibs_cbs = (valor_liquido * aliq_total / Decimal('100')).quantize(Decimal('0.01'))
            
            prod['ibs_cbs'] = ibs_cbs
            prod['valor_liquido'] = valor_liquido
            prod['total_tributos'] = total_tributos
        
        produtos_saidas = sorted(produtos_saidas_dict.values(), key=lambda x: (-x['valor_total'], x['codigo']))
        
        # ========================================
        # BUSCAR PRODUTOS SALVOS VIA API (se não houver do SPED)
        # ========================================
        produtos_saidas_api = []
        if not produtos_saidas:
            # Buscar produtos salvos via API
            produtos_api_db = ProdutoSaidaAPI.objects.filter(
                empresa_id=empresa_id,
                periodo_inicial=periodo_inicial,
                periodo_final=periodo_final
            )
            
            if produtos_api_db.exists():
                # Agrupar produtos por código + NCM
                produtos_api_dict = {}
                for prod in produtos_api_db:
                    chave = f"{prod.codigo}_{prod.ncm}"
                    if chave not in produtos_api_dict:
                        produtos_api_dict[chave] = {
                            'chave_nfe': prod.chave_nfe,
                            'codigo': prod.codigo,
                            'descricao': prod.descricao,
                            'ncm': prod.ncm,
                            'cfop': prod.cfop,
                            'quantidade': Decimal('0'),
                            'valor_total': Decimal('0'),
                            'icms': Decimal('0'),
                            'icms_st': Decimal('0'),
                            'ipi': Decimal('0'),
                            'pis': Decimal('0'),
                            'cofins': Decimal('0'),
                        }
                    produtos_api_dict[chave]['quantidade'] += prod.quantidade or Decimal('0')
                    produtos_api_dict[chave]['valor_total'] += prod.valor_total or Decimal('0')
                    produtos_api_dict[chave]['icms'] += prod.icms_valor or Decimal('0')
                    produtos_api_dict[chave]['icms_st'] += prod.icms_st_valor or Decimal('0')
                    produtos_api_dict[chave]['ipi'] += prod.ipi_valor or Decimal('0')
                    produtos_api_dict[chave]['pis'] += prod.pis_valor or Decimal('0')
                    produtos_api_dict[chave]['cofins'] += prod.cofins_valor or Decimal('0')
                
                # Calcular campos adicionais para produtos da API
                for chave, prod in produtos_api_dict.items():
                    total_tributos = prod['icms'] + prod['icms_st'] + prod['ipi'] + prod['pis'] + prod['cofins']
                    valor_liquido = prod['valor_total'] - total_tributos
                    aliq_total = aliquota_ibs + aliquota_cbs + aliquota_is
                    ibs_cbs = (valor_liquido * aliq_total / Decimal('100')).quantize(Decimal('0.01'))
                    total_reforma = valor_liquido + ibs_cbs
                    dif_total = ibs_cbs - total_tributos
                    
                    # Valores unitários
                    qtd = prod['quantidade'] if prod['quantidade'] > 0 else Decimal('1')
                    valor_bruto_unit = (prod['valor_total'] / qtd).quantize(Decimal('0.01'))
                    valor_liq_unit = (valor_liquido / qtd).quantize(Decimal('0.01'))
                    ibs_cbs_unit = (ibs_cbs / qtd).quantize(Decimal('0.01'))
                    total_reforma_unit = (total_reforma / qtd).quantize(Decimal('0.01'))
                    dif_unit = (dif_total / qtd).quantize(Decimal('0.01'))
                    
                    prod['ibs_cbs'] = ibs_cbs
                    prod['valor_liquido'] = valor_liquido
                    prod['total_tributos'] = total_tributos
                    prod['total_reforma'] = total_reforma
                    prod['dif_total'] = dif_total
                    prod['valor_bruto_unit'] = valor_bruto_unit
                    prod['valor_liq_unit'] = valor_liq_unit
                    prod['ibs_cbs_unit'] = ibs_cbs_unit
                    prod['total_reforma_unit'] = total_reforma_unit
                    prod['dif_unit'] = dif_unit
                    prod['aliq_ibs_cbs'] = aliq_total
                    prod['perc_reducao'] = Decimal('0')
                    
                    # Formatar para exibição
                    prod['quantidade_fmt'] = f"{prod['quantidade']:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                    prod['perc_reducao_fmt'] = f"{prod['perc_reducao']:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                    prod['aliq_ibs_cbs_fmt'] = f"{aliq_total:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                    prod['valor_bruto_unit_fmt'] = f"{valor_bruto_unit:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                    prod['valor_bruto_fmt'] = f"{prod['valor_total']:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                    prod['icms_fmt'] = f"{prod['icms']:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                    prod['icms_st_fmt'] = f"{prod['icms_st']:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                    prod['ipi_fmt'] = f"{prod['ipi']:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                    prod['pis_fmt'] = f"{prod['pis']:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                    prod['cofins_fmt'] = f"{prod['cofins']:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                    prod['total_tributos_fmt'] = f"{total_tributos:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                    prod['valor_liq_unit_fmt'] = f"{valor_liq_unit:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                    prod['valor_liq_fmt'] = f"{valor_liquido:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                    prod['ibs_cbs_unit_fmt'] = f"{ibs_cbs_unit:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                    prod['ibs_cbs_fmt'] = f"{ibs_cbs:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                    prod['total_reforma_unit_fmt'] = f"{total_reforma_unit:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                    prod['total_reforma_fmt'] = f"{total_reforma:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                    prod['dif_unit_fmt'] = f"{dif_unit:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                    prod['dif_total_fmt'] = f"{dif_total:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                
                produtos_saidas_api = sorted(produtos_api_dict.values(), key=lambda x: (-x['valor_total'], x['codigo']))
                
                # Usar produtos da API se não houver do SPED
                produtos_saidas = produtos_saidas_api
                
                # Recalcular totais de saídas com dados da API
                total_valor_saidas = sum(p['valor_total'] for p in produtos_saidas)
                total_icms_saidas = sum(p['icms'] for p in produtos_saidas)
                total_pis_saidas = sum(p['pis'] for p in produtos_saidas)
                total_cofins_saidas = sum(p['cofins'] for p in produtos_saidas)
                total_ibs_cbs_saidas = sum(p['ibs_cbs'] for p in produtos_saidas)
                debitos_saidas = total_icms_saidas + total_pis_saidas + total_cofins_saidas
                venda_liquida = total_valor_saidas - debitos_saidas
                carga_saidas = (debitos_saidas / total_valor_saidas * 100) if total_valor_saidas else Decimal('0')
                carga_saidas_reforma = (total_ibs_cbs_saidas / venda_liquida * 100) if venda_liquida else Decimal('0')
        
# Totais ENTRADAS
        total_valor_entradas = sum(p['valor_total'] for p in produtos_entradas)
        total_icms_entradas = sum(p['icms'] for p in produtos_entradas)
        total_icms_st_entradas = sum(p['icms_st'] for p in produtos_entradas)
        total_ipi_entradas = sum(p.get('ipi', Decimal('0')) for p in produtos_entradas)
        total_pis_entradas = sum(p['pis'] for p in produtos_entradas)
        total_cofins_entradas = sum(p['cofins'] for p in produtos_entradas)
        total_ibs_cbs_entradas = sum(p['ibs_cbs'] for p in produtos_entradas)
        creditos_entradas = total_icms_entradas + total_icms_st_entradas + total_ipi_entradas + total_pis_entradas + total_cofins_entradas
        compra_liquida = total_valor_entradas - creditos_entradas
        carga_entradas = (creditos_entradas / total_valor_entradas * 100) if total_valor_entradas else Decimal('0')
        carga_entradas_reforma = (total_ibs_cbs_entradas / compra_liquida * 100) if compra_liquida else Decimal('0')
        
        # Totais detalhados para tfoot de Produtos Adquiridos
        total_tributos_entradas = sum(p['total_tributos'] for p in produtos_entradas)
        total_valor_liquido_entradas = sum(p['valor_liquido'] for p in produtos_entradas)
        total_valor_bruto_unit_entradas = sum(p['valor_bruto_unit'] for p in produtos_entradas)
        total_valor_liq_unit_entradas = sum(p['valor_liq_unit'] for p in produtos_entradas)
        total_ibs_cbs_unit_entradas = sum(p['ibs_cbs_unit'] for p in produtos_entradas)
        total_reforma_entradas = sum(p['total_reforma'] for p in produtos_entradas)
        total_reforma_unit_entradas = sum(p['total_reforma_unit'] for p in produtos_entradas)
        total_dif_total_entradas = sum(p['dif_total'] for p in produtos_entradas)
        total_dif_unit_entradas = sum(p['dif_unit'] for p in produtos_entradas)
        
        # Totais SAÍDAS
        total_valor_saidas = sum(p['valor_total'] for p in produtos_saidas)
        total_icms_saidas = sum(p['icms'] for p in produtos_saidas)
        total_pis_saidas = sum(p['pis'] for p in produtos_saidas)
        total_cofins_saidas = sum(p['cofins'] for p in produtos_saidas)
        total_ibs_cbs_saidas = sum(p['ibs_cbs'] for p in produtos_saidas)
        debitos_saidas = total_icms_saidas + total_pis_saidas + total_cofins_saidas
        venda_liquida = total_valor_saidas - debitos_saidas
        carga_saidas = (debitos_saidas / total_valor_saidas * 100) if total_valor_saidas else Decimal('0')
        carga_saidas_reforma = (total_ibs_cbs_saidas / venda_liquida * 100) if venda_liquida else Decimal('0')
        
        # Apuração geral
        resultado_atual = debitos_saidas - creditos_entradas
        resultado_reforma = total_ibs_cbs_saidas - total_ibs_cbs_entradas
        carga_atual = (resultado_atual / total_valor_saidas * 100) if total_valor_saidas else Decimal('0')
        carga_reforma = (resultado_reforma / venda_liquida * 100) if venda_liquida else Decimal('0')
        
# Função para formatar valores monetários no padrão brasileiro
        def formatar_valor(valor):
            return f"{valor:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
        
        # Atualizar contexto
        context.update({
            # Produtos
            'produtos_entradas': produtos_entradas,
            'produtos_saidas': produtos_saidas,
            
            # Totais Entradas
            'total_valor_entradas': total_valor_entradas,
            'total_icms_entradas': total_icms_entradas,
            'total_pis_entradas': total_pis_entradas,
            'total_cofins_entradas': total_cofins_entradas,
            'total_ibs_cbs_entradas': total_ibs_cbs_entradas,
            'compra_bruta': total_valor_entradas,
            'compra_bruta_fmt': formatar_valor(total_valor_entradas),
            'creditos_entradas': creditos_entradas,
            'creditos_entradas_fmt': formatar_valor(creditos_entradas),
            'compra_liquida': compra_liquida,
            'compra_liquida_fmt': formatar_valor(compra_liquida),
            'carga_entradas': f"{carga_entradas:.2f}".replace('.', ','),
            'compra_liquida_reforma': compra_liquida,
            'compra_liquida_reforma_fmt': formatar_valor(compra_liquida),
            'creditos_ibs_cbs': total_ibs_cbs_entradas,
            'creditos_ibs_cbs_fmt': formatar_valor(total_ibs_cbs_entradas),
            'compra_total_reforma': compra_liquida + total_ibs_cbs_entradas,
            'compra_total_reforma_fmt': formatar_valor(compra_liquida + total_ibs_cbs_entradas),
            'carga_entradas_reforma': f"{carga_entradas_reforma:.2f}".replace('.', ','),
            
            # Totais tfoot Entradas
            'total_icms_st_entradas': total_icms_st_entradas,
            'total_ipi_entradas': total_ipi_entradas,
            'total_tributos_entradas': total_tributos_entradas,
            'total_valor_liquido_entradas': total_valor_liquido_entradas,
            'total_valor_bruto_unit_entradas': total_valor_bruto_unit_entradas,
            'total_valor_liq_unit_entradas': total_valor_liq_unit_entradas,
            'total_ibs_cbs_unit_entradas': total_ibs_cbs_unit_entradas,
            'total_reforma_entradas': total_reforma_entradas,
            'total_reforma_unit_entradas': total_reforma_unit_entradas,
            'total_dif_total_entradas': total_dif_total_entradas,
            'total_dif_unit_entradas': total_dif_unit_entradas,
            
            # Totais Saídas
            'total_valor_saidas': total_valor_saidas,
            'total_icms_saidas': total_icms_saidas,
            'total_pis_saidas': total_pis_saidas,
            'total_cofins_saidas': total_cofins_saidas,
            'total_ibs_cbs_saidas': total_ibs_cbs_saidas,
            'venda_bruta': total_valor_saidas,
            'debitos_saidas': debitos_saidas,
            'venda_liquida': venda_liquida,
            'carga_saidas': f"{carga_saidas:.2f}".replace('.', ','),
            'venda_liquida_reforma': venda_liquida,
            'debitos_ibs_cbs': total_ibs_cbs_saidas,
            'venda_total_reforma': venda_liquida + total_ibs_cbs_saidas,
            'carga_saidas_reforma': f"{carga_saidas_reforma:.2f}".replace('.', ','),
            
            # Apuração Resumo
            'debitos_atual': debitos_saidas,
            'creditos_atual': creditos_entradas,
            'resultado_atual': resultado_atual,
            'carga_atual': f"{carga_atual:.2f}".replace('.', ','),
            'debitos_reforma': total_ibs_cbs_saidas,
            'creditos_reforma': total_ibs_cbs_entradas,
            'resultado_reforma': resultado_reforma,
            'carga_reforma': f"{carga_reforma:.2f}".replace('.', ','),
            
            # Dados para gráficos
            'icms_entradas': float(total_icms_entradas),
            'pis_entradas': float(total_pis_entradas),
            'cofins_entradas': float(total_cofins_entradas),
            'ibs_cbs_entradas': float(total_ibs_cbs_entradas),
            'icms_saidas': float(total_icms_saidas),
            'pis_saidas': float(total_pis_saidas),
            'cofins_saidas': float(total_cofins_saidas),
            'ibs_cbs_saidas': float(total_ibs_cbs_saidas),
            'carga_atual_compras': float(carga_entradas),
            'carga_reforma_compras': float(carga_entradas_reforma),
            'carga_atual_vendas': float(carga_saidas),
            'carga_reforma_vendas': float(carga_saidas_reforma),
            
            # Fornecedores
            'fornecedores_entradas': fornecedores_entradas,
            'total_fornecedores_valor_bruto': total_fornecedores_valor_bruto,
            'total_fornecedores_icms': total_fornecedores_icms,
            'total_fornecedores_icms_st': total_fornecedores_icms_st,
            'total_fornecedores_ipi': total_fornecedores_ipi,
            'total_fornecedores_iss': total_fornecedores_iss,
            'total_fornecedores_pis': total_fornecedores_pis,
            'total_fornecedores_cofins': total_fornecedores_cofins,
            'total_fornecedores_tributos': total_fornecedores_tributos,
            'total_fornecedores_liquido': total_fornecedores_liquido,
            'total_fornecedores_ibs_cbs': total_fornecedores_ibs_cbs,
            'total_fornecedores_reforma': total_fornecedores_reforma,
            
            # CFOPs de Entrada
            'cfops_entradas': cfops_entradas,
            'total_cfops_valor_bruto': total_cfops_valor_bruto,
            'total_cfops_icms': total_cfops_icms,
            'total_cfops_icms_st': total_cfops_icms_st,
            'total_cfops_ipi': total_cfops_ipi,
            'total_cfops_pis': total_cfops_pis,
            'total_cfops_cofins': total_cfops_cofins,
            'total_cfops_tributos': total_cfops_tributos,
            'total_cfops_liquido': total_cfops_liquido,
            
            # UFs de Entrada
            'ufs_entradas': ufs_entradas,
            'total_ufs_valor_bruto': total_ufs_valor_bruto,
            'total_ufs_icms': total_ufs_icms,
            'total_ufs_icms_st': total_ufs_icms_st,
            'total_ufs_ipi': total_ufs_ipi,
            'total_ufs_iss': total_ufs_iss,
            'total_ufs_pis': total_ufs_pis,
            'total_ufs_cofins': total_ufs_cofins,
            'total_ufs_tributos': total_ufs_tributos,
            'total_ufs_liquido': total_ufs_liquido,
            
# Ajustes Apuração ICMS
            'ajustes_icms': ajustes_icms,
            'total_ajustes_icms': total_ajustes_icms,
            
            # Ajustes Manuais ICMS
            'ajustes_manuais_icms': ajustes_manuais_icms,
            'total_ajustes_manuais_icms': total_ajustes_manuais_icms,
        })
    
    return render(request, 'dashboards/relatorio_fiscal.html', context)
@login_required
def api_periodos(request):
    empresa_id = request.GET.get('empresa')
    if not empresa_id:
        return JsonResponse({'periodos': []})
    
    from apps.sped.models import Registro0000
    periodos = Registro0000.objects.filter(
        empresa_id=empresa_id,
        processado=True
    ).values_list('periodo', flat=True).distinct().order_by('periodo')
    
    return JsonResponse({'periodos': [p.strftime('%Y-%m') for p in periodos]})


@login_required
def api_chaves_processadas(request):
    """Retorna lista de chaves NF-e já processadas e salvas no banco"""
    empresa_id = request.GET.get('empresa')
    periodo_inicial = request.GET.get('periodo_inicial')
    periodo_final = request.GET.get('periodo_final')
    
    if not empresa_id or not periodo_inicial or not periodo_final:
        return JsonResponse({'chaves_processadas': []})
    
    # Buscar chaves já processadas no banco
    chaves_salvas = ProdutoSaidaAPI.objects.filter(
        empresa_id=empresa_id,
        periodo_inicial=periodo_inicial,
        periodo_final=periodo_final
    ).values_list('chave_nfe', flat=True).distinct()
    
    return JsonResponse({
        'chaves_processadas': list(chaves_salvas),
        'total_processadas': len(chaves_salvas)
    })


@login_required
def api_listar_chaves_saida(request):
    """Retorna lista de chaves de acesso das notas de saída do período selecionado"""
    from apps.sped.models import Registro0000, RegistroC100
    
    empresa_id = request.GET.get('empresa')
    periodo_inicial = request.GET.get('periodo_inicial')
    periodo_final = request.GET.get('periodo_final')
    
    if not empresa_id or not periodo_inicial or not periodo_final:
        return JsonResponse({'success': False, 'message': 'Parâmetros incompletos', 'chaves': []})
    
    registros_0000 = Registro0000.objects.filter(
        empresa_id=empresa_id,
        processado=True,
        periodo__gte=periodo_inicial + '-01',
        periodo__lte=periodo_final + '-28'
    )
    
# Buscar TODOS os documentos C100 de saída (ind_oper='1') com chave válida
    documentos_saida = RegistroC100.objects.filter(
        registro_0000__in=registros_0000,
        ind_oper='1'
    ).exclude(chv_nfe__isnull=True).exclude(chv_nfe='')
    
    chaves_nfe = []  # NF-e para buscar
    chaves_nfce = []  # NFC-e ignoradas
    
    for doc in documentos_saida:
        if doc.chv_nfe and len(doc.chv_nfe) >= 44:
            modelo = doc.cod_mod or '55'
            dados_doc = {
                'chave_nfe': doc.chv_nfe,
                'numero': doc.num_doc,
                'modelo': modelo,
                'modelo_desc': 'NF-e' if modelo == '55' else 'NFC-e',
                'valor': float(doc.vl_doc) if doc.vl_doc else 0,
                'data': doc.dt_doc.strftime('%d/%m/%Y') if doc.dt_doc else '',
            }
            
            if modelo == '55':
                chaves_nfe.append(dados_doc)
            else:
                chaves_nfce.append(dados_doc)
    
    return JsonResponse({
        'success': True,
        'total': len(chaves_nfe),
        'total_nfe': len(chaves_nfe),
        'total_nfce': len(chaves_nfce),
        'chaves': chaves_nfe,  # Apenas NF-e para buscar
        'nfce_ignoradas': chaves_nfce  # NFC-e para relatório
    })


from django.views.decorators.csrf import csrf_exempt

import logging
logger = logging.getLogger('dashboards.api')

@csrf_exempt
@login_required
def api_buscar_xml_produto(request):
    """Busca XML de uma NFe via API MeuDanfe e extrai dados dos produtos"""
    import requests
    import base64
    import xml.etree.ElementTree as ET
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Método não permitido'}, status=405)
    
    import json
    data = json.loads(request.body)
    chave_acesso = data.get('chave_acesso', '')
    
    if not chave_acesso:
        logger.warning("[API-XML] Chave de acesso não informada")
        return JsonResponse({'success': False, 'message': 'Chave de acesso não informada'})
    
    # Limpar chave
    chave_acesso = ''.join(filter(str.isdigit, chave_acesso))
    num_nota = chave_acesso[25:34].lstrip('0') if len(chave_acesso) >= 34 else '?'
    
    if len(chave_acesso) != 44:
        logger.warning(f"[API-XML] NF {num_nota} - Chave inválida: {len(chave_acesso)} dígitos")
        return JsonResponse({'success': False, 'message': f'Chave inválida: {len(chave_acesso)} dígitos'})
    
    logger.info(f"[API-XML] Buscando NF {num_nota} - Chave: {chave_acesso[:20]}...")
    
    # API Key do MeuDanfe
    api_key = '167af6d1-a260-4a6b-bc48-729064cc4efd'
    
    headers = {
        'Api-Key': api_key,
        'Content-Type': 'application/json'
    }
    
    # Identificar modelo pela chave (posição 20-21 da chave = modelo)
    modelo = chave_acesso[20:22] if len(chave_acesso) >= 22 else '55'
    logger.info(f"[API-XML] NF {num_nota} - Modelo: {modelo} ({'NFC-e' if modelo == '65' else 'NF-e'})")
    
    try:
        # Passo 1: Adicionar chave para busca
        # NFC-e (65) usa endpoint diferente
        if modelo == '65':
            url_add = f'https://api.meudanfe.com.br/v2/nfce/add/{chave_acesso}'
        else:
            url_add = f'https://api.meudanfe.com.br/v2/fd/add/{chave_acesso}'
        response = requests.put(url_add, headers=headers, timeout=30)
        
        if response.status_code == 200:
            data_resp = response.json()
            search_status = data_resp.get('SearchStatus') or data_resp.get('searchStatus') or data_resp.get('status')
            logger.info(f"[API-XML] NF {num_nota} - Status: {search_status}")
            
            if search_status in ['FOUND', 'Found', 'found', 'OK', 'ok', 'SUCCESS', 'success']:
# Passo 2: Buscar XML
                if modelo == '65':
                    url_xml = f'https://api.meudanfe.com.br/v2/nfce/get/xml/{chave_acesso}'
                else:
                    url_xml = f'https://api.meudanfe.com.br/v2/fd/get/xml/{chave_acesso}'
                response_xml = requests.get(url_xml, headers=headers, timeout=30)
                
                if response_xml.status_code == 200:
                    data_xml = response_xml.json()
                    xml_content = None
                    
                    if data_xml.get('data') and data_xml.get('data').startswith('<?xml'):
                        xml_content = data_xml.get('data')
                    elif data_xml.get('Base64') or data_xml.get('base64'):
                        xml_base64 = data_xml.get('Base64') or data_xml.get('base64')
                        xml_content = base64.b64decode(xml_base64).decode('utf-8')
                    
                    if xml_content:
                        # Extrair dados dos produtos
                        produtos = extrair_produtos_xml(xml_content, chave_acesso)
                        logger.info(f"[API-XML] NF {num_nota} - ✅ SUCESSO! {len(produtos)} produtos extraídos")
                        return JsonResponse({
                            'success': True,
                            'chave': chave_acesso,
                            'produtos': produtos,
                            'qtd_produtos': len(produtos)
                        })

                    else:
                        return JsonResponse({'success': False, 'message': 'XML não retornado pela API'})
                else:
                    return JsonResponse({'success': False, 'message': f'Erro ao buscar XML: {response_xml.status_code}'})
            else:
                logger.warning(f"[API-XML] NF {num_nota} - Não encontrada: {search_status}")
                return JsonResponse({'success': False, 'message': f'NFe não encontrada: {search_status}'})
        elif response.status_code == 402:
            logger.error(f"[API-XML] NF {num_nota} - Saldo insuficiente")
            return JsonResponse({'success': False, 'message': 'Saldo insuficiente na API'})
        elif response.status_code == 401:
            logger.error(f"[API-XML] NF {num_nota} - API Key inválida")
            return JsonResponse({'success': False, 'message': 'API Key inválida'})
        else:
            logger.error(f"[API-XML] NF {num_nota} - Erro HTTP: {response.status_code}")
            return JsonResponse({'success': False, 'message': f'Erro na API: {response.status_code}'})
    
    except Exception as e:
        logger.error(f"[API-XML] NF {num_nota} - Exceção: {str(e)}")
        return JsonResponse({'success': False, 'message': f'Erro: {str(e)}'})


def extrair_produtos_xml(xml_content, chave_acesso):
    """Extrai dados dos produtos de um XML de NFe"""
    import xml.etree.ElementTree as ET
    from decimal import Decimal
    
    produtos = []
    
    try:
        root = ET.fromstring(xml_content)
        ns = {'nfe': 'http://www.portalfiscal.inf.br/nfe'}
        
        # Buscar itens (det)
        for det in root.findall('.//nfe:det', ns):
            prod = det.find('nfe:prod', ns)
            imposto = det.find('nfe:imposto', ns)
            
            if prod is None:
                continue
            
            item = {
                'chave_nfe': chave_acesso,
                'num_item': det.get('nItem', ''),
                'codigo': prod.findtext('nfe:cProd', '', ns),
                'ean': prod.findtext('nfe:cEAN', '', ns),
                'descricao': prod.findtext('nfe:xProd', '', ns),
                'ncm': prod.findtext('nfe:NCM', '', ns),
                'cest': prod.findtext('nfe:CEST', '', ns),
                'cfop': prod.findtext('nfe:CFOP', '', ns),
                'unidade': prod.findtext('nfe:uCom', '', ns),
                'quantidade': float(prod.findtext('nfe:qCom', '0', ns) or 0),
                'valor_unitario': float(prod.findtext('nfe:vUnCom', '0', ns) or 0),
                'valor_total': float(prod.findtext('nfe:vProd', '0', ns) or 0),
            }
            
# Impostos
            if imposto:
                # ICMS
                icms = imposto.find('.//nfe:ICMS', ns)
                if icms:
                    for icms_tipo in icms:
                        item['icms_cst'] = icms_tipo.findtext('nfe:CST', '', ns) or icms_tipo.findtext('nfe:CSOSN', '', ns)
                        item['icms_aliq'] = float(icms_tipo.findtext('nfe:pICMS', '0', ns) or 0)
                        item['icms_valor'] = float(icms_tipo.findtext('nfe:vICMS', '0', ns) or 0)
                        # ICMS ST
                        item['icms_st_valor'] = float(icms_tipo.findtext('nfe:vICMSST', '0', ns) or 0)
                        item['icms_st_bc'] = float(icms_tipo.findtext('nfe:vBCST', '0', ns) or 0)
                        break
                
                # IPI
                ipi = imposto.find('.//nfe:IPI', ns)
                if ipi:
                    for ipi_tipo in ipi:
                        item['ipi_cst'] = ipi_tipo.findtext('nfe:CST', '', ns)
                        item['ipi_aliq'] = float(ipi_tipo.findtext('nfe:pIPI', '0', ns) or 0)
                        item['ipi_valor'] = float(ipi_tipo.findtext('nfe:vIPI', '0', ns) or 0)
                        break
                else:
                    item['ipi_cst'] = ''
                    item['ipi_aliq'] = 0
                    item['ipi_valor'] = 0
                
                # PIS
                pis = imposto.find('.//nfe:PIS', ns)
                if pis:
                    for pis_tipo in pis:
                        item['pis_cst'] = pis_tipo.findtext('nfe:CST', '', ns)
                        item['pis_aliq'] = float(pis_tipo.findtext('nfe:pPIS', '0', ns) or 0)
                        item['pis_valor'] = float(pis_tipo.findtext('nfe:vPIS', '0', ns) or 0)
                        break
                
                # COFINS
                cofins = imposto.find('.//nfe:COFINS', ns)
                if cofins:
                    for cofins_tipo in cofins:
                        item['cofins_cst'] = cofins_tipo.findtext('nfe:CST', '', ns)
                        item['cofins_aliq'] = float(cofins_tipo.findtext('nfe:pCOFINS', '0', ns) or 0)
                        item['cofins_valor'] = float(cofins_tipo.findtext('nfe:vCOFINS', '0', ns) or 0)
                        break
            
            produtos.append(item)
    
    except Exception as e:
        print(f"Erro ao extrair produtos do XML: {e}")
    
    return produtos
@login_required
def debug_sped(request):
    from django.http import JsonResponse
    from apps.sped.models import Registro0000
    from apps.empresa.models import Empresa
    
    # Todas as empresas
    empresas = list(Empresa.objects.all().values('id', 'cnpj_cpf', 'razao_social', 'ativo'))
    
    # Todos os registros SPED
    registros = list(Registro0000.objects.all().values(
        'id', 'empresa_id', 'empresa__cnpj_cpf', 'tipo', 'periodo', 'processado', 'created_at'
    ))
    
    return JsonResponse({
        'empresas': empresas,
        'registros_sped': registros
    }, json_dumps_params={'default': str})
@csrf_exempt
@login_required
def api_exportar_relatorio_erros(request):
    """Exporta relatório de NF-e com erro e NFC-e ignoradas para Excel"""
    import json
    from io import BytesIO
    from django.http import HttpResponse
    
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return JsonResponse({'success': False, 'message': 'openpyxl não instalado'})
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Método não permitido'}, status=405)
    
    data = json.loads(request.body)
    nfe_erros = data.get('nfe_erros', [])
    nfce_ignoradas = data.get('nfce_ignoradas', [])
    
    wb = openpyxl.Workbook()
    
    # Estilos
    header_font = Font(bold=True, color='FFFFFF')
    header_fill_erro = PatternFill(start_color='DC2626', end_color='DC2626', fill_type='solid')
    header_fill_nfce = PatternFill(start_color='F59E0B', end_color='F59E0B', fill_type='solid')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Aba NF-e com Erro
    ws_erros = wb.active
    ws_erros.title = 'NF-e com Erro'
    headers_erro = ['Número', 'Chave de Acesso', 'Data', 'Valor', 'Erro']
    for col, header in enumerate(headers_erro, 1):
        cell = ws_erros.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill_erro
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    for row, nfe in enumerate(nfe_erros, 2):
        ws_erros.cell(row=row, column=1, value=nfe.get('numero', '')).border = border
        ws_erros.cell(row=row, column=2, value=nfe.get('chave_nfe', '')).border = border
        ws_erros.cell(row=row, column=3, value=nfe.get('data', '')).border = border
        ws_erros.cell(row=row, column=4, value=nfe.get('valor', 0)).border = border
        ws_erros.cell(row=row, column=5, value=nfe.get('erro', '')).border = border
    
    # Ajustar largura das colunas
    ws_erros.column_dimensions['A'].width = 12
    ws_erros.column_dimensions['B'].width = 50
    ws_erros.column_dimensions['C'].width = 12
    ws_erros.column_dimensions['D'].width = 15
    ws_erros.column_dimensions['E'].width = 40
    
    # Aba NFC-e Ignoradas
    ws_nfce = wb.create_sheet('NFC-e Ignoradas')
    headers_nfce = ['Número', 'Chave de Acesso', 'Data', 'Valor', 'Motivo']
    for col, header in enumerate(headers_nfce, 1):
        cell = ws_nfce.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill_nfce
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    for row, nfce in enumerate(nfce_ignoradas, 2):
        ws_nfce.cell(row=row, column=1, value=nfce.get('numero', '')).border = border
        ws_nfce.cell(row=row, column=2, value=nfce.get('chave_nfe', '')).border = border
        ws_nfce.cell(row=row, column=3, value=nfce.get('data', '')).border = border
        ws_nfce.cell(row=row, column=4, value=nfce.get('valor', 0)).border = border
        ws_nfce.cell(row=row, column=5, value=nfce.get('motivo', '')).border = border
    
    # Ajustar largura das colunas
    ws_nfce.column_dimensions['A'].width = 12
    ws_nfce.column_dimensions['B'].width = 50
    ws_nfce.column_dimensions['C'].width = 12
    ws_nfce.column_dimensions['D'].width = 15
    ws_nfce.column_dimensions['E'].width = 40
    
    # Salvar em memória
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=relatorio_erros_nfe.xlsx'
    return response
@csrf_exempt
@login_required
def api_salvar_produtos_api(request):
    """Salva os produtos buscados via API no banco de dados"""
    import json
    from decimal import Decimal
    from .models import ProdutoSaidaAPI
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Método não permitido'}, status=405)
    
    try:
        data = json.loads(request.body)
        empresa_id = data.get('empresa_id')
        periodo_inicial = data.get('periodo_inicial')
        periodo_final = data.get('periodo_final')
        produtos = data.get('produtos', [])
        
        if not empresa_id or not periodo_inicial or not periodo_final:
            return JsonResponse({'success': False, 'message': 'Parâmetros obrigatórios não informados'})
        
        if not produtos:
            return JsonResponse({'success': False, 'message': 'Nenhum produto para salvar'})
        
        # Remover produtos anteriores do mesmo período
        ProdutoSaidaAPI.objects.filter(
            empresa_id=empresa_id,
            periodo_inicial=periodo_inicial,
            periodo_final=periodo_final
        ).delete()
        
        # Salvar novos produtos
        produtos_criados = 0
        for prod in produtos:
            ProdutoSaidaAPI.objects.create(
                empresa_id=empresa_id,
                periodo_inicial=periodo_inicial,
                periodo_final=periodo_final,
                chave_nfe=prod.get('chave_nfe', ''),
                codigo=prod.get('codigo', ''),
                descricao=prod.get('descricao', ''),
                ncm=prod.get('ncm', ''),
                cfop=prod.get('cfop', ''),
                unidade=prod.get('unidade', ''),
                quantidade=Decimal(str(prod.get('quantidade', 0) or 0)),
                valor_unitario=Decimal(str(prod.get('valor_unitario', 0) or 0)),
                valor_total=Decimal(str(prod.get('valor_total', 0) or 0)),
                icms_cst=prod.get('icms_cst', ''),
                icms_aliq=Decimal(str(prod.get('icms_aliq', 0) or 0)),
                icms_valor=Decimal(str(prod.get('icms_valor', 0) or 0)),
                icms_st_valor=Decimal(str(prod.get('icms_st_valor', 0) or 0)),
                ipi_cst=prod.get('ipi_cst', ''),
                ipi_aliq=Decimal(str(prod.get('ipi_aliq', 0) or 0)),
                ipi_valor=Decimal(str(prod.get('ipi_valor', 0) or 0)),
                pis_cst=prod.get('pis_cst', ''),
                pis_aliq=Decimal(str(prod.get('pis_aliq', 0) or 0)),
                pis_valor=Decimal(str(prod.get('pis_valor', 0) or 0)),
                cofins_cst=prod.get('cofins_cst', ''),
                cofins_aliq=Decimal(str(prod.get('cofins_aliq', 0) or 0)),
                cofins_valor=Decimal(str(prod.get('cofins_valor', 0) or 0)),
            )
            produtos_criados += 1
        
        return JsonResponse({
            'success': True,
            'message': f'{produtos_criados} produtos salvos com sucesso',
            'qtd_salvos': produtos_criados
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Erro ao salvar: {str(e)}'})
@csrf_exempt
@login_required
def api_importar_ajustes_manuais_icms(request):
    """Importa ajustes manuais ICMS via planilha (JSON)"""
    import json
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Método não permitido'}, status=405)
    
    try:
        data = json.loads(request.body)
        empresa_id = data.get('empresa_id')
        periodo_inicial = data.get('periodo_inicial')
        periodo_final = data.get('periodo_final')
        ajustes = data.get('ajustes', [])
        
        if not empresa_id or not periodo_inicial or not periodo_final:
            return JsonResponse({'success': False, 'message': 'Parâmetros obrigatórios ausentes'})
        
        criados = 0
        for aj in ajustes:
            AjusteManualICMS.objects.create(
                empresa_id=empresa_id,
                periodo_inicial=periodo_inicial,
                periodo_final=periodo_final,
                codigo=aj.get('codigo', ''),
                descricao=aj.get('descricao', ''),
                valor=Decimal(str(aj.get('valor', 0)).replace(',', '.')),
            )
            criados += 1
        
        return JsonResponse({'success': True, 'message': f'{criados} ajuste(s) importado(s) com sucesso'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@csrf_exempt
@login_required
def api_excluir_ajustes_manuais_icms(request):
    """Exclui todos os ajustes manuais ICMS de um período"""
    import json
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Método não permitido'}, status=405)
    
    try:
        data = json.loads(request.body)
        empresa_id = data.get('empresa_id')
        periodo_inicial = data.get('periodo_inicial')
        periodo_final = data.get('periodo_final')
        
        deletados = AjusteManualICMS.objects.filter(
            empresa_id=empresa_id,
            periodo_inicial=periodo_inicial,
            periodo_final=periodo_final
        ).delete()
        
        return JsonResponse({'success': True, 'message': f'{deletados[0]} ajuste(s) excluído(s)'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@csrf_exempt
@login_required
def api_excluir_ajuste_manual_icms(request, ajuste_id):
    """Exclui um ajuste manual ICMS individual"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Método não permitido'}, status=405)
    
    try:
        ajuste = AjusteManualICMS.objects.get(id=ajuste_id)
        ajuste.delete()
        return JsonResponse({'success': True, 'message': 'Ajuste excluído com sucesso'})
    except AjusteManualICMS.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Ajuste não encontrado'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})    